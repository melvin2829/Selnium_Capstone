import openpyxl

def read_excel_data(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    data = []
    for row in range(2, sheet.max_row + 1):
        row_data = [sheet.cell(row, col).value for col in range(1, sheet.max_column + 1)]
        data.append(row_data)
    return data
