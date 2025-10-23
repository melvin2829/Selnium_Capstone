import openpyxl
from openpyxl import Workbook
import os

# File path
file_path = "data/test_data.xlsx"

# Create folder if it doesn't exist
if not os.path.exists("data"):
    os.makedirs("data")

# Create a new workbook
wb = Workbook()

# --- Registration Sheet ---
if "registration" in wb.sheetnames:
    ws_reg = wb["registration"]
else:
    ws_reg = wb.active
    ws_reg.title = "registration"

ws_reg.append(["FirstName","LastName","Email","Telephone","Password","ConfirmPassword","Subscribe"])
ws_reg.append(["John","Doe","john1@example.com","9876543210","Pass@123","Pass@123","Yes"])
ws_reg.append(["Jane","Smith","jane1@example.com","9876543211","Pass@123","Pass@123","No"])

# --- Login Sheet ---
ws_login = wb.create_sheet("login")
ws_login.append(["Username","Password","ExpectedResult"])
ws_login.append(["john1@example.com","Pass@123","Login Successful"])
ws_login.append(["john1@example.com","wrong123","Login Failed"])

# --- Search Sheet ---
ws_search = wb.create_sheet("search")
ws_search.append(["Product"])
ws_search.append(["iPhone"])
ws_search.append(["MacBook"])

# Save workbook
wb.save(file_path)
print(f"Test data added successfully to {file_path}")
